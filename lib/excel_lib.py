from openpyxl import Workbook
from openpyxl import load_workbook
	
def excel_file_init(header):
	wb = Workbook()
	ws = wb.active
	ws.append(header)
	return wb,ws

def add_row_to_excel_file(ws,row):
	ws.append(row)

def save_excel_file(wb,file):
	wb.save("%s.xlsx"%file)

def excel_to_file(sheet):
	rows = sheet.nrows	
	for row in range(rows):
		yield sheet.row_values(row)

def load_excel_file(filename):
	return load_workbook(filename)